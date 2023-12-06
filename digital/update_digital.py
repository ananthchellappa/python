#!/usr/bin/env python3


import os
import sys
import shutil
import datetime

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


SRC_SHEET_NAME = "Digital Pins"
DEFAULT_OUTPUT_FILE = "digital.sv"
PIN_COLUMN_NAME = "Pin"
TYPE_COLUMN_NAME = "Type"
DEFAULT_COLUMN_NAME = "Default"
DESCRIPTION_COLUMN_NAME = "Description"


def header_squeeze(h):
    return h.lower().strip().replace(" ", "").replace(".", "").replace("-", "").replace("_", "").replace("/", "")


def header_key(h):
    return h.lower().replace(" ", "_")


MANDATORY_HEADERS = {
    PIN_COLUMN_NAME,
    TYPE_COLUMN_NAME,
}

OPTIONAL_HEADERS = {
    DEFAULT_COLUMN_NAME,
    DESCRIPTION_COLUMN_NAME,
}

# INPUT_TYPE_MATCHES = [
#     "input",
#     "ipnut",
#     "inpt",
#     "inp",
#     "in",
# ]

# OUTPUT_TYPE_MATCHES = [
#     "output",
#     "otput",
#     "outpt",
#     "outp",
#     "out",
# ]


def read_pin_details_from_worksheet(excel_file, excel_ws, sv_file_format):
    try:
        excel_cells = excel_ws["A1:{}{}".format(get_column_letter(excel_ws.max_column), excel_ws.max_row)]
        excel_data = []
        for row in excel_cells:
            excel_data.append([str(c.value) if c.value is not None else "" for c in row])
    except Exception as e:
        return [], "Error reading excel data: {}: {}".format(type(e), e)

    if sv_file_format is not True:
        sv_file_format = False

    header_indices = {}

    for header_index, row in enumerate(excel_data):
        squeezed_row = [header_squeeze(v) for v in row]
        if all([header_squeeze(h) in squeezed_row for h in MANDATORY_HEADERS]):
            header_indices = {header_key(h): squeezed_row.index(header_squeeze(h)) for h in MANDATORY_HEADERS}
            for oh in OPTIONAL_HEADERS:
                if header_squeeze(oh) in squeezed_row:
                    header_indices[header_key(oh)] = squeezed_row.index(header_squeeze(oh))
            break

    if not header_indices:
        return [], "Failed to detect all mandatory headers"

    pin_details = []

    for index, row in enumerate(excel_data[header_index + 1:]):
        pin_desc_str = row[header_indices[header_key(PIN_COLUMN_NAME)]].strip()
        type_str = row[header_indices[header_key(TYPE_COLUMN_NAME)]].strip()
        if header_key(DEFAULT_COLUMN_NAME) in header_indices:
            default_value_str = row[header_indices[header_key(DEFAULT_COLUMN_NAME)]].strip()
        else:
            default_value_str = ""
        if header_key(DESCRIPTION_COLUMN_NAME) in header_indices:
            description_str = row[header_indices[header_key(DESCRIPTION_COLUMN_NAME)]].strip()
        else:
            description_str = ""

        if pin_desc_str and type_str:
            bus_bits = None
            pin_name = pin_desc_str.split(">", 1)[0].split("<", 1)[0].strip()
            if "<" in pin_desc_str and ">" in pin_desc_str:
                bus_bits = pin_desc_str.rsplit(">", 1)[0].split("<", 1)[1].strip()

            # if type_str.lower() in INPUT_TYPE_MATCHES:
            #     pin_type = "input"
            # elif type_str.lower() in OUTPUT_TYPE_MATCHES:
            #     pin_type = "output"
            if type_str.lower().startswith("i"):
                pin_type = "input"
            elif type_str.lower().startswith("o"):
                pin_type = "output"
            else:
                error = "Invalid PIN Type {} in input file {}".format(type_str, excel_file)
                print(error, file=sys.stderr)
                return [], error

            comment = ""
            if default_value_str:
                comment = default_value_str
            if description_str:
                if default_value_str:
                    comment += ", "
                comment += description_str

            pin_details.append(dict(
                pin_type=pin_type,
                logic_word=sv_file_format,
                bus_bits=bus_bits,
                pin_name=pin_name,
                possible_last_line=False,
                comment=comment,
            ))
        else:
            pin_details.append(dict())

    return pin_details, None


def parse_code_line(line_index, line, sv_file_format):

    stripped_line = line.strip()
    if not stripped_line:
        return {}, None

    comment = ""
    if "//" in stripped_line:
        comment = stripped_line.split("//", 1)[1].lstrip()
    code_line = stripped_line.split("//", 1)[0]
    if not code_line:
        return {}, None

    segments = [l.strip() for l in code_line.split() if l.strip()]

    pin_type = ""

    segment_index = 0

    if len(segments) > segment_index and segments[segment_index] in ["input", "output"]:
        pin_type = segments[segment_index]
        segment_index += 1
    else:
        error = "Missing pin type in line {}: {}".format(line_index + 1, line)
        print(error, file=sys.stderr)
        return {}, error

    logic_word = False
    if len(segments) > segment_index + 1 and segments[segment_index] in ["logic"]:
        logic_word = True
        segment_index += 1

    if sv_file_format != logic_word:
        if sv_file_format:
            print("Missing 'logic' word in line number {}: {}".format(line_index + 1, line.rstrip()))
        else:
            print("Redundant 'logic' word in line number {}: {}".format(line_index + 1, line.rstrip()))

    bus_bits = None
    if len(segments) > segment_index and segments[segment_index].startswith("[") and segments[segment_index].endswith("]"):
        bus_bits = segments[segment_index][1:-1]
        segment_index += 1

    if len(segments) > segment_index:
        if segments[segment_index].endswith(","):
            pin_name = segments[segment_index][:-1].strip()
            possible_last_line = False
        else:
            pin_name = segments[segment_index]
            possible_last_line = True
    else:
        error = "Missing pin name in line {}: {}".format(line_index + 1, line)
        print(error, file=sys.stderr)
        return {}, error

    return dict(
        pin_type=pin_type,
        logic_word=logic_word,
        bus_bits=bus_bits,
        pin_name=pin_name,
        possible_last_line=possible_last_line,
        comment=comment,
        original_line=line
    ), None


def read_pin_details_from_text_file(text_file, sv_file_format):
    pin_details = []
    expected_module_name = os.path.basename(text_file).rsplit(".", 1)[0]
    try:
        with open(text_file, encoding="utf-8") as fd:
            lines = fd.readlines()

            module_read = False
            module_open_read = False
            module_close_read = False

            for line_index, line in enumerate(lines):
                result_line = [line_index]

                if not module_read:
                    code_line = line.split("//")[0]
                    if "module" in code_line:
                        module_read = True
                        result_line.append("module")
                        module_name = code_line.split("module", 1)[1].split()[0]
                        if module_name != expected_module_name:
                            error = "Module name extracted '{}' does not match with expected '{}'".format(module_name, expected_module_name)
                            print(error, file=sys.stderr)
                            line = line.split("module", 1)[0] + "module" + line.split("module", 1)[1].replace(module_name, expected_module_name)
                        if "(" in code_line:
                            module_open_read = True
                            result_line.append("module-open")

                    result_line.append(line)
                    pin_details.append(result_line)
                    continue

                if not module_open_read:
                    code_line = line.split("//")[0]
                    if "(" in code_line:
                        result_line.append("module-open")
                        module_open_read = True
                    result_line.append(line)
                    pin_details.append(result_line)
                    continue

                if not module_close_read:
                    code_line = line.split("//")[0]
                    if ");" in code_line.replace(" ", "").replace("\t", ""):
                        result_line.append("module-close")
                        module_close_read = True
                        result_line.append(line)
                        pin_details.append(result_line)
                        continue
                    else:
                        pin_line_dict, error = parse_code_line(line_index, line, sv_file_format)
                        if not error:
                            if pin_line_dict:
                                result_line.append(pin_line_dict)
                                pin_details.append(result_line)
                            else:
                                result_line.append(line)
                                pin_details.append(result_line)
                        else:
                            print("Error processing line:", error, file=sys.stderr)
                            result_line.append(line)
                            pin_details.append(result_line)
                else:
                    result_line.append(line)
                    pin_details.append(result_line)
                    continue

            if module_read and module_open_read and module_close_read:
                return pin_details, None

            error = "Missing module block in file: {}".format(text_file)
            print(error, file=sys.stderr)
            return [], error
    except Exception as e:
        error = "Exception reading PIN details from file: {}: {}: {}".format(text_file, type(e), e)
        print(error, file=sys.stderr)
        return [], error


def merge_pin_details(input_excel, existing_file_content, module_name, pin_details, verbose_mode, sv_file_format):
    error = None
    output_lines = []

    added = []
    updated = []
    deleted = []
    missing_comments = []

    if not existing_file_content:
        output_lines.append("module {}\n".format(module_name))
        output_lines.append("(\n")
        pin_count = len(pin_details)
        if pin_details:
            output_lines.append("    // adding ports from {}\n".format(os.path.basename(input_excel)))

        for pin_index, pin_info in enumerate(pin_details):

            if verbose_mode:
                line_comment = " // {}".format(pin_info["comment"]) if pin_info["comment"] else ""
            else:
                line_comment = ""

            if verbose_mode and not pin_info["comment"]:
                pass
                # missing_comments.append(pin_info["pin_name"])

            line = "    {}{}{} {}{}{}\n".format(
                pin_info["pin_type"],
                " logic" if sv_file_format else "",
                " [{}]".format(pin_info["bus_bits"]) if pin_info["bus_bits"] else "",
                pin_info["pin_name"],
                "," if pin_index < pin_count - 1 else "",
                line_comment
            )
            output_lines.append(line)
        output_lines.append(");\n")
    else:
        module_opened = False
        module_closed = False
        last_saved_pin_line = None
        retained_pins = set()

        excel_dict = {l["pin_name"]: l for l in pin_details}
        for file_line in existing_file_content:
            line_index = file_line.pop(0)
            line_content = file_line.pop()
            comment_updated = False

            if module_opened and not module_closed:
                if isinstance(line_content, dict):
                    pin_name = line_content["pin_name"]
                    if pin_name not in excel_dict:
                        deleted.append(pin_name)
                        if output_lines and output_lines[-1].strip().startswith("//") and pin_name in output_lines[-1]:
                            output_lines.pop()
                    else:
                        retained_pins.add(pin_name)
                        line_updated = False
                        comment = line_content["comment"]

                        if verbose_mode:
                            if comment != excel_dict[pin_name]["comment"]:
                                if comment and not excel_dict[pin_name]["comment"]:
                                    missing_comments.append(pin_name)
                                comment = excel_dict[pin_name]["comment"]
                                # line_updated = True
                                comment_updated = True

                        pin_type = line_content["pin_type"]
                        if pin_type != excel_dict[pin_name]["pin_type"]:
                            pin_type = excel_dict[pin_name]["pin_type"]
                            line_updated = True

                        logic_word = line_content["logic_word"]
                        if logic_word != excel_dict[pin_name]["logic_word"]:
                            logic_word = excel_dict[pin_name]["logic_word"]
                            line_updated = True

                        bus_bits = line_content["bus_bits"]
                        if bus_bits != excel_dict[pin_name]["bus_bits"]:
                            bus_bits = excel_dict[pin_name]["bus_bits"]
                            line_updated = True

                        if not line_updated and not comment_updated:
                            output_lines.append(line_content["original_line"])
                        else:
                            updated_line = "    {}{}{} {},{}\n".format(
                                pin_type,
                                " logic" if logic_word else "",
                                " [{}]".format(bus_bits) if bus_bits else "",
                                pin_name,
                                " // {}".format(comment) if comment else ""
                            )
                            output_lines.append(updated_line)
                            if line_updated:
                                updated.append(pin_name)
                        last_saved_pin_line = len(output_lines) - 1, pin_name + ",", pin_name

                    continue

            if not module_opened and "module-open" in file_line:
                module_opened = True

            if not module_closed and "module-close" in file_line:
                module_closed = True
                add_comment_added = False

                for pin_info in pin_details:
                    pin_name = pin_info["pin_name"]
                    if pin_name in retained_pins:
                        continue

                    if verbose_mode:
                        line_comment = " // {}".format(pin_info["comment"]) if pin_info["comment"] else ""
                    else:
                        line_comment = ""

                    if verbose_mode and not pin_info["comment"]:
                        pass
                        # missing_comments.append(pin_name)

                    line = "    {}{}{} {},{}\n".format(
                        pin_info["pin_type"],
                        " logic" if sv_file_format else "",
                        " [{}]".format(pin_info["bus_bits"]) if pin_info["bus_bits"] else "",
                        pin_name,
                        line_comment,
                    )

                    if not add_comment_added:
                        add_comment_added = True
                        output_lines.append("    // adding ports from {}\n".format(os.path.basename(input_excel)))
                    added.append(pin_name)
                    output_lines.append(line)
                    last_saved_pin_line = len(output_lines) - 1, pin_name + ",", pin_name

                if last_saved_pin_line is not None:
                    output_lines[last_saved_pin_line[0]] = output_lines[last_saved_pin_line[0]].replace(last_saved_pin_line[1], last_saved_pin_line[2])

                output_lines.append(line_content)

                if deleted:
                    now = datetime.datetime.now()
                    delete_summary = "// {} {} {} deleted".format(
                        now.strftime("%m/%d/%Y %H:%M:%S"),
                        "pin" if len(deleted) == 1 else "pins",
                        deleted[0] if len(deleted) == 1 else " and ".join([", ".join(deleted[:-1]), deleted[-1]])
                    )
                    output_lines.append(delete_summary)
                continue

            output_lines.append(line_content)

    return dict(
        output_lines=output_lines,
        added=added,
        updated=updated,
        deleted=deleted,
        missing_comments=missing_comments,
    ), error


def main(excel_file, output_file, verbose_mode=False):
    try:
        wb = load_workbook(filename=excel_file)
        ws = wb[SRC_SHEET_NAME]
    except Exception as e:
        return False, "Error opening input excel file {}: {}:{}".format(excel_file, type(e), e)

    sv_file_format = False
    if output_file.lower().endswith(".sv"):
        sv_file_format = True

    pin_details, error = read_pin_details_from_worksheet(excel_file, ws, sv_file_format)
    if not pin_details:
        return False, "Failed to read register map from: {} Error: {}".format(excel_file, error)

    existing_file_content = []
    module_name = os.path.basename(output_file).rsplit(".", 1)[0]
    if os.path.isfile(output_file):
        backup_file = output_file + ".bak"
        print("Backing up existing output file '{}' to '{}'".format(output_file, backup_file))
        shutil.copy(output_file, backup_file)
        existing_file_content, error = read_pin_details_from_text_file(output_file, sv_file_format=sv_file_format)

    result, error = merge_pin_details(excel_file, existing_file_content, module_name, pin_details, verbose_mode, sv_file_format)
    if not error:

        output_lines = result["output_lines"]
        added = result.get("added", [])
        updated = result.get("updated", [])
        deleted = result.get("deleted", [])
        missing_comments = result.get("missing_comments", [])

        with open(output_file, "w", encoding="utf-8") as fd:
            fd.writelines(output_lines)

        if added:
            print("Added pin lines:", added)
        if updated:
            print("Updated pin lines:", updated)
        if deleted:
            print("Deleted pin lines:", deleted)
        if missing_comments:
            print("Pins missing comments in input Excel:", missing_comments)

    return True, None


if __name__ == '__main__':

    # sys.argv += ["-v", "source.xlsx", "digital.sv"]
    # sys.argv += ["-v", "source.xlsx", "test.v"]
    # sys.argv += ["source_2.xlsx", "-v", "digital.sv"]

    verbose_mode = False

    if "-v" in sys.argv:
        sys.argv.remove("-v")
        verbose_mode = True

    if "-verbose" in sys.argv:
        sys.argv.remove("-verbose")
        verbose_mode = True

    if "--verbose" in sys.argv:
        sys.argv.remove("--verbose")
        verbose_mode = True

    if len(sys.argv) < 2:
        print("Usage: {} [-v/--verbose] <excel-file> [output-file]".format(os.path.basename(__file__)), file=sys.stderr)
        sys.exit(1)

    excel_file = sys.argv[1]

    output_file = DEFAULT_OUTPUT_FILE
    if len(sys.argv) > 2:
        output_file = sys.argv[2]

    error = None

    if not excel_file or not os.path.isfile(excel_file):
        error = "Invalid input excel file: {}".format(os.path.realpath(excel_file) if excel_file else excel_file)

    if error is None:
        status, error = main(excel_file, output_file=output_file, verbose_mode=verbose_mode)
        if not status:
            print("Failed to process file {}:".format(excel_file), error, file=sys.stderr)
    else:
        print("Failed with error:", error, file=sys.stderr)
