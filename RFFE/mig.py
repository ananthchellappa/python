#!/usr/bin/env python3

# NOTE : supports HEX addresses (3rd field) ONLY - so run your script through this if necessary : 
# perl -p -e 's/^\s*([^,]+,[^,]+),(\d+)/sprintf("$1,0x%02x",oct("0b$2"))/e unless /^\s*[^,]+,[^,]+,0[xX]/;'

import os
import sys

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


REGISTER_MAP_SHEET = "Register Map Detail"

MANDATORY_HEADERS = [
    "Register Address", "Data Bits", "Bit Field Name", "Default"
]


def header_squeeze(h):
    return h.lower().strip().replace(" ", "").replace(".", "").replace("-", "").replace("_", "").replace("/", "")


def header_key(h):
    return h.lower().replace(" ", "_")


def parse_register_address(register_address):
    try:
        if register_address.lower().startswith("0x"):
            return dict(
                string=register_address.strip(),
                integer=int(register_address.strip(), 16)
            )
        else:
            return dict(
                string=register_address.strip(),
                integer=int(register_address.strip(), 2)
            )
    except Exception:
        raise ValueError("Invalid register address: {}".format(register_address))


def parse_data_bits(data_bits):
    try:
        name = data_bits.split("[", 1)[0].strip()
        if not name:
            raise ValueError("Invalid data_bits name: {}".format(data_bits))
    except Exception as e:
        raise e

    if "[" in data_bits and "]" in data_bits:
        bit_positions = data_bits.split("[")[-1].split("]")[0]
        start_bit = bit_positions.split(":", 1)[0]
        if start_bit.isdigit():
            start_bit = int(start_bit)
        else:
            raise ValueError("Nondigit start bit: {}".format(data_bits))
        if ":" in bit_positions:
            end_bit = bit_positions.split(":", 1)[-1]
            if end_bit.isdigit():
                end_bit = int(end_bit)
            else:
                raise ValueError("Nondigit end bit: {}".format(data_bits))
        else:
            end_bit = start_bit
    else:
         start_bit = 7
         end_bit = 0

    return dict(name=name, start_bit=start_bit, end_bit=end_bit)


def parse_bit_field_name(bit_field_name):
    try:
        name = bit_field_name.split("[", 1)[0].strip().replace(" ", "_").lower()
        original_name = bit_field_name.split("[", 1)[0].strip()
        if not name:
            raise ValueError("Invalid bit_field name: {}".format(bit_field_name))
    except Exception as e:
        raise e

    if "[" in bit_field_name and "]" in bit_field_name:
        bit_positions = bit_field_name.split("[")[-1].split("]")[0]
        start_bit = bit_positions.split(":", 1)[0]
        if start_bit.isdigit():
            start_bit = int(start_bit)
        else:
            raise ValueError("Nondigit start bit: {}".format(bit_field_name))
        if ":" in bit_positions:
            end_bit = bit_positions.split(":", 1)[-1]
            if end_bit.isdigit():
                end_bit = int(end_bit)
            else:
                raise ValueError("Nondigit end bit: {}".format(bit_field_name))
        else:
            end_bit = start_bit
    else:
         start_bit = 7
         end_bit = 0

    if name.lower() == "unused":
        unused = True
    else:
        unused = False

    return dict(name=name, original_name=original_name, start_bit=start_bit, end_bit=end_bit, unused=unused)


def parse_default_value(default_value):
    try:
        length, int_value = int(default_value.lower().split('b', 1)[0]), int(default_value.lower().split('b', 1)[1], 2)
    except Exception:
        raise ValueError("Invalid default_value: {}".format(default_value))

    value = f'{int_value:08b}'
    if length > len(value):
        value = '0' * (length - len(value)) + value
    elif length < len(value):
        value = value[-length:]

    try:
        bits = []
        if length == len(value):
            for b in value:
                if int(b) in [0, 1]:
                    bits.append(int(b))
                else:
                    raise Exception(value)
        else:
            raise Exception(value)
    except Exception:
        raise ValueError("Invalid default_value: {}".format(default_value))

    return dict(bits=bits)


def read_register_map(excel_ws):

    try:
        excel_cells = excel_ws["A1:{}{}".format(get_column_letter(excel_ws.max_column), excel_ws.max_row)]
        excel_data = []
        for row in excel_cells:
            excel_data.append([str(c.value) if c.value is not None else "" for c in row])
    except Exception as e:
        return [], "Error reading excel data: {}: {}".format(type(e), e)

    header_indices = {}

    for header_index, row in enumerate(excel_data):
        squeezed_row = [header_squeeze(v) for v in row]
        if all([header_squeeze(h) in squeezed_row for h in MANDATORY_HEADERS]):
            header_indices = {header_key(h): squeezed_row.index(header_squeeze(h)) for h in MANDATORY_HEADERS}
            break

    if not header_indices:
        return [], "Failed to detect all mandatory headers"

    register_map = []
    for index, row in enumerate(excel_data[header_index + 1:]):
        register_address = row[header_indices["register_address"]]
        data_bits = row[header_indices["data_bits"]]
        # bit_field_name = row[header_indices["bit_field_name"]]
        bit_field_name = row[header_indices[MANDATORY_HEADERS[2].replace(" ","_").lower()]]
        default_value = row[header_indices["default"]]

        if not register_address or not data_bits or not bit_field_name or not bit_field_name:
            continue

        try:
            register_map.append(dict(
                register_address=parse_register_address(register_address),
                data_bits=parse_data_bits(data_bits),
                bit_field_name=parse_bit_field_name(bit_field_name),
                default_value=parse_default_value(default_value),
            ))
        except ValueError as e:
            print("Failed to parse row {}".format(header_index + 1 + index), file=sys.stderr)

    return register_map, None


def translate_line(old_reg_address_dict, old_bitfieds_dict, new_reg_address_dict, new_bitfieds_dict, old_line, previous_writes=None):
    print("Processing line:", old_line)
    if "//" in old_line:
        existing_comment = old_line.split("//", 1)[1].strip()
    else:
        existing_comment = None

    old_line = old_line.split("//")[0].strip()
    command_segments = [s.strip() for s in old_line.strip().split(",")]
    if len(command_segments) >= 4:

        previous_writes = previous_writes or {}

        # write_command = command_segments[0]
        usid_str = command_segments[1]
        address_str = command_segments[2]
        data_str = command_segments[3]

        if data_str.lower().startswith("0x"):
            data_int = int(data_str.strip(), 16)
            data_bits = [int(c) for c in f'{data_int:08b}']
        else:
            data_bits = [int(c) for c in data_str if c in ['1', '0']]

        if len(data_bits) < 8:
            data_bits = [0] * (8 - len(data_bits)) + data_bits
        address = integer=int(address_str.strip(), 16)

        old_register_details = old_reg_address_dict.get(address)
        if not old_register_details:
            print("{}: Could not find address {} in old register".format(old_line, address_str))
            return [old_line], previous_writes
        else:
            old_bitfield_data = [r.get("bit_field_name", {}) for r in old_register_details]

            old_bitfield_names = [b.get("name") for b in old_bitfield_data]
            old_bitfield_unused = [b.get("unused", True) for b in old_bitfield_data]
            old_bitfield_defaults = [r.get("default_value", {}).get("bits") for r in old_register_details]
            old_reg_bits = [(r.get("data_bits", {}).get("start_bit"), r.get("data_bits", {}).get("end_bit")) for r in old_register_details]
            old_reg_data_segments = [data_bits[::-1][b[1]:b[0] + 1][::-1] for b in old_reg_bits]

            modified_regs = {}

            for i in range(len(old_bitfield_names)):

                bitfield_name = old_bitfield_names[i]
                bitfield_unused = old_bitfield_unused[i]
                old_reg_segment_value = old_reg_data_segments[i]
                old_default_value = old_bitfield_defaults[i]

                if not bitfield_unused:
                    new_bitfield_data = new_bitfieds_dict.get(bitfield_name)
                    if new_bitfield_data:
                        for bdata in new_bitfield_data:
                            new_address = bdata.get("register_address", {}).get("integer")
                            new_default_value = bdata.get("default_value", {}).get("bits")
                            new_register_details = new_reg_address_dict.get(new_address)
                            new_reg_bitfield_names = [r.get("bit_field_name", {}).get("name") for r in new_register_details]
                            new_reg_bits = [(r.get("data_bits", {}).get("start_bit"), r.get("data_bits", {}).get("end_bit")) for r in new_register_details]

                            if bitfield_name in new_reg_bitfield_names:
                                new_reg_segment_index = new_reg_bitfield_names.index(bitfield_name)
                                if old_default_value != old_reg_segment_value:
                                    if new_address not in modified_regs:
                                        modified_regs[new_address] = []
                                    modified_regs[new_address].append((bitfield_name, new_reg_bits[new_reg_segment_index], old_reg_segment_value))

            mod_lines = []
            for reg in modified_regs:
                mod_data = modified_regs[reg]

                mod_functions = []
                mod_values = []
                size_changed = set()

                for entry in mod_data:
                    mod_functions.append(entry[0])
                    mod_bits_len = entry[1][0] - entry[1][1] + 1
                    mod_bits_values = entry[2]
                    while mod_bits_len > len(mod_bits_values):
                        mod_bits_values = [0] + mod_bits_values
                        size_changed.add(entry[0])
                    while mod_bits_values and mod_bits_len < len(mod_bits_values):
                        mod_bits_values = mod_bits_values[1:]
                        size_changed.add(entry[0])
                    mod_values.append(mod_bits_values)

                def cmp_databits(r):
                    startbit = r.get("data_bits", {}).get("start_bit", 0)
                    endbit = r.get("data_bits", {}).get("end_bit", 0)
                    return (8 - startbit) * 1000 + (8 - endbit)

                mod_register_details = sorted(new_reg_address_dict.get(reg), key=lambda r: cmp_databits(r))
                mod_data_segments = []
                mod_data_functions = []

                original_names = []
                size_warning = False

                for seg in mod_register_details:
                    seg_fn = seg.get("bit_field_name", {}).get("name")
                    original_name = seg.get("bit_field_name", {}).get("original_name")

                    if seg_fn in mod_functions:
                        if seg_fn in size_changed:
                            original_names.append("{} has different size".format(original_name))
                            size_warning = True
                        else:
                            original_names.append(original_name)
                        mod_data_segments.append(mod_values[mod_functions.index(seg_fn)])
                        mod_data_functions.append(seg_fn)
                        previous_writes[seg_fn] = address, mod_values[mod_functions.index(seg_fn)]
                    elif seg_fn in previous_writes:
                        old_saved_address, saved_segment = previous_writes[seg_fn]
                        if old_saved_address != address:
                            mod_data_segments.append(saved_segment)
                        else:
                            mod_data_segments.append(seg.get("default_value", {}).get("bits"))
                    else:
                        mod_data_segments.append(seg.get("default_value", {}).get("bits"))

                new_address = new_reg_address_dict.get(reg)[0].get("register_address").get("integer")

                if 32 <= new_address:
                    new_write_command = "ew"
                else:
                    new_write_command = "w"

                if not size_warning and existing_comment:
                    comment_part = " // {}".format(existing_comment)
                else:
                    comment_part = " // {}{}".format("WARNING - " if size_warning else "", ",".join(original_names)) if original_names else ""

                mod_lines.append("{},{},{},{}{}".format(
                    new_write_command,
                    usid_str,
                    new_reg_address_dict.get(reg)[0].get("register_address").get("string"),
                    "_".join(["".join([str(i) for i in l]) for l in mod_data_segments]),
                    comment_part,
                ))
                print("Translated line:", mod_lines[-1])

            if mod_lines:
                return mod_lines, previous_writes
            else:
                return [old_line], previous_writes
    else:
        print("Could not find enough segments in line:", old_line, file=sys.stderr)
        return [old_line], previous_writes



def main(old_excel, new_excel, old_cmd_text, new_cmd_text=None, previous_writes=None):

    try:
        old_wb = load_workbook(filename=old_excel)
        old_ws = old_wb[REGISTER_MAP_SHEET]
    except Exception as e:
        return False, "Error opening input excel file {}: {}:{}".format(old_excel, type(e), e)

    try:
        new_wb = load_workbook(filename=new_excel)
        new_ws = new_wb[REGISTER_MAP_SHEET]
    except Exception as e:
        return False, "Error opening input excel file {}: {}:{}".format(new_excel, type(e), e)

    try:
        with open(old_cmd_text, encoding="utf-8") as fd:
            old_lines = [l for l in fd.readlines()]
            stripped_lines = [l.strip() for l in old_lines]
    except Exception as e:
        return False, "Error opening input text file {}: {}:{}".format(old_cmd_text, type(e), e)

    if not new_cmd_text:
        input_folder = os.path.dirname(old_cmd_text)
        input_filename = os.path.basename(old_cmd_text)
        output_folder = os.path.join(input_folder, "migrated")
        os.makedirs(output_folder, exist_ok=True)
        new_cmd_text = os.path.join(output_folder, input_filename)

    print("\n\nInput file:", old_cmd_text)
    print("Output file:", new_cmd_text)

    previous_writes = previous_writes or {}

    old_register_map, error = read_register_map(old_ws)
    if not old_register_map:
        return False, "Failed to read register map from: {} Error: {}".format(old_excel, error)

    new_register_map, error = read_register_map(new_ws)
    if not new_register_map:
        return False, "Failed to read register map from: {} Error: {}".format(new_excel, error)

    old_reg_address_dict = {}
    old_bitfieds_dict = {}

    for row_dict in old_register_map:
        reg_address_key = row_dict["register_address"]["integer"]
        if reg_address_key not in old_reg_address_dict:
            old_reg_address_dict[reg_address_key] = [row_dict]
        else:
            old_reg_address_dict[reg_address_key].append(row_dict)

        bit_field_key = row_dict["bit_field_name"]["name"]
        if bit_field_key not in old_bitfieds_dict:
            old_bitfieds_dict[bit_field_key] = [row_dict]
        else:
            old_bitfieds_dict[bit_field_key].append(row_dict)

    new_reg_address_dict = {}
    new_bitfieds_dict = {}

    for row_dict in new_register_map:
        reg_address_key = row_dict["register_address"]["integer"]
        if reg_address_key not in new_reg_address_dict:
            new_reg_address_dict[reg_address_key] = [row_dict]
        else:
            new_reg_address_dict[reg_address_key].append(row_dict)

        bit_field_key = row_dict["bit_field_name"]["name"]
        if bit_field_key not in new_bitfieds_dict:
            new_bitfieds_dict[bit_field_key] = [row_dict]
        else:
            new_bitfieds_dict[bit_field_key].append(row_dict)

    new_lines = []
    for line_index, old_line in enumerate(old_lines):
        if stripped_lines[line_index].lower().startswith("ew") or stripped_lines[line_index].lower().startswith("w"):
            line_result, previous_writes = translate_line(old_reg_address_dict, old_bitfieds_dict, new_reg_address_dict, new_bitfieds_dict, old_line, previous_writes=previous_writes)
            for lr in line_result:
                new_lines.append(lr)
            if old_line.endswith("\n"):
                new_lines[-1] += "\n"
        else:
            new_lines.append(old_line)

    new_lines = [l if l.endswith("\n") else l + "\n" for l in new_lines]

    with open(new_cmd_text, "w", encoding="utf-8") as fd:
        fd.write("".join(new_lines))

    return True, None


if __name__ == '__main__':

    # print(sys.argv)
    # sys.exit(0)

    # sys.argv += ["oldREG.xlsx", "newREG.xlsx", "ex9.txt"]
    # sys.argv += ["old_0x59.xlsx", "new_0x59.xlsx", "ex9.txt"]

    if len(sys.argv) < 4:
        print("Usage: {} <old-excel-file> <new-excel-file> <old-cmd-text>".format(os.path.basename(__file__)), file=sys.stderr)
        sys.exit(1)

    old_excel = sys.argv[1]
    new_excel = sys.argv[2]
    old_cmd_texts = sys.argv[3:]

    error = None

    if not old_excel or not os.path.isfile(old_excel):
        error = "Invalid input excel file: {}".format(os.path.realpath(old_excel) if old_excel else old_excel)
    elif not new_excel or not os.path.isfile(new_excel):
        error = "Invalid input excel file: {}".format(os.path.realpath(new_excel) if new_excel else new_excel)

    if error is None:
        for old_cmd_text in old_cmd_texts:
            if not old_cmd_text or not os.path.isfile(old_cmd_text):
                error = "Invalid input text file: {}".format(os.path.realpath(old_cmd_text) if old_cmd_text else old_cmd_text)
                print("Failed to process file {}:".format(old_cmd_text), error, file=sys.stderr)
            else:
                new_cmd_text = os.path.join(os.path.dirname(os.path.realpath(old_cmd_text)), "new_" + os.path.basename(old_cmd_text))
                status, error = main(old_excel, new_excel, old_cmd_text)
                if not status:
                    print("Failed to process file {}:".format(old_cmd_text), error, file=sys.stderr)

    else:
        print("Failed with error:", error, file=sys.stderr)
